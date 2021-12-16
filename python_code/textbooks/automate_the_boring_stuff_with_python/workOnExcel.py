import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())

print(wb.get_sheet_by_name('Sheet3'))
sheet3 = wb.get_sheet_by_name('Sheet3')
type(sheet3)

print(sheet3.title)
# anothersheet = wb.get_active_sheet()
# print(anothersheet)

sheet1 = wb.get_sheet_by_name('Sheet1')
print(sheet1)
print(sheet1['A1'].value)

c = sheet1['B1']
print(c.value)
string = 'Row ' + str(c.row) + ', column'+ str(c.column) + ' is ' + c.value
print(string)

string2 = f'Cell {c.coordinate} is {c.value}'
print(string2)
print(sheet1['C1'].value)

print(sheet1.cell(row=1, column=2))
print(sheet1.cell(row=1, column=2).value)

for i in range(1, 8, 2):
	print(i, sheet1.cell(row=i, column=2).value)

# print(sheet1.get_highest_row())

# from openpyxl.cell import get_column_letter, column_index_from_string
# print(get_column_letter(1))


	# print(tuple(sheet1['A1':'C3']))
	# for rowofcellobjs in sheet1['A1':'C3']:
	# 	for cellobj in rowofcellobjs:
	# 		print(cellobj.coordinate, cellobj.value)
	# 	print(' end of row '.center(20, '#'))


for cellobj in sheet1.columns[1]:
	print(cellobj.value)



