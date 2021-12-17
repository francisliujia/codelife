import re 
import openpyxl

print("enter the absolute path of the spreadsheet you wish to split:")
file_path = input()

path_regx = re.compile(r"(./*)(.*)(\.xlsx)$")
path_split = path_regx.search(file_path)
path = path_split.group(1)
name = path_split.group(2)

wb = openpyxl.load_workbook(file_path)
sheet = wb.active

column_num = 1
for column in range(1, sheet.max_column + 1):
	column_data = []
	for cell in range(1, sheet.max_row + 1):
		if sheet.cell(row=cell, column=column_num).value != None:
			column_data.append(sheet.cell(row=cell, column=column_num).value)

	file = open('/' + path + 'col-' + str(column_num) + '-' + '.txt'. 'w')
	for line in column_data:
		file.write(line + '\n')
	file.close()

	column_num += 1



