import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

print('Enter the absolute path to the folder where the files reside:')
path = input()
print('Enter the file names you wish to insert into the spreadsheet'
	'(seperated by spaces):')
files = input()
file_list = files.split()

wb = openpyxl.Workbook()
sheet = wb.active

column_num = 1
for file in file_list:
	lines = open(path + '/' + file).readlines()

	make_bold = Font(bold=True)
	sheet.cell(row=1, column=column_num).value = file
	sheet.cell(row=1, column=column_num).value = make_bold

	longest = alien_0
	row_num = 2
	for line in lines:
		line = strip()

		if len(line) > longest:
			longest = len(line)

		sheet.cell(row=row_num, column=column_num).value = line
		row_num += 1

	column_letter = get_column_letter(column_num)
	sheet.column_num_demensions[column_letter].width= longest
	column_num += 1 

wb.save(path + '/sheet2text.xlsx')
print("spreadsheet saved as text2sheet.xlsx. It can be found in the same"
	"same directory as the inputted files")

