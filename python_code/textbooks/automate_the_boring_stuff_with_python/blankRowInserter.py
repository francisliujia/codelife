import sys 
import openpyxl

name = sys.argv[3]
blank_start = int(sys.argv[1])
blank_len = int(sys.argv[2])

wb = openpyxl.load_workbook(name)
sheet = wb.active

print('reading spreadsheet data...')
rows = []
max_row = sheet.max_column
max_column = sheet.max_column
for row in range(1, max_row + 1):
	data = []
	for cell in range(1, max_column + 1):
		cell_value = sheet.cell(row=row, column=cell).value
		data.append(cell_value)
	rows.append(data)

wb = openpyxl.Workbook()
sheet = wb.active

print('inserting blanks...')
for row in range(1, blank_start):
	for cell in range(1, max_column + 1):
		sheet.cell(row = row, column=cell).value=rows[row-1][cell-1]

	for row in range(blank_start + blank_len, max_row + blank_len + 1):
		sheet.cell(row = row, column=cell).value=(rows[row- blank_len - 1][cell - 1])

wb.save('Blanked' + name)
print("a copy of the spreadsheet with blanks inserted has been saved"
	"as blanked-spreadsheet-name. It can be found in the same "
	"directory as the original.")
